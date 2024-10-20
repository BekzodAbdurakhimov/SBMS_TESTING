from openpyxl import Workbook, load_workbook
import time
from datetime import datetime
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from config.config import RATE_PLAN_NAME_1, RATE_PLAN_NAME_2

LOG_FILE = "../results/logs.txt"
EXCEL_FILE = "результаты_смены_тп.xlsx"


def log_step(message):
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_message = f"[{current_time}] {message}\n"
    print(log_message.strip())
    with open(LOG_FILE, "a") as file:
        file.write(log_message)


def test_basic_operations(driver):
    global wb, ws

    # Попробуйте открыть существующий файл или создать новый
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        # ws['A1'] = "СМЕНА ТП"
        # ws['B1'] = "Баланс до"
        # ws['C1'] = "Баланс после"
        # ws['D1'] = "Результат смены ТП"

    driver.get("https://sbms.ucell/ps/sbms/shell.html")
    wait = WebDriverWait(driver, 120)

    log_step(' ======== Смена ТП ========')

    # # Проверка кнопки
    # detail_button_check = wait.until(
    #     EC.visibility_of_element_located((By.ID, "details-button"))
    # )
    # detail_button_check.click()
    #
    # go_to_link = wait.until(
    #     EC.visibility_of_element_located((By.ID, "proceed-link"))
    # )
    # go_to_link.click()
    # login_text_locator = (By.CSS_SELECTOR, '.login-caption > span')
    # login_locator_check = wait.until(EC.visibility_of_element_located(login_text_locator))
    # assert login_locator_check.is_displayed(), "Не найден локатор входа"
    #
    # login_input_locator = (By.CSS_SELECTOR, "input.sbms-textbox[name='user'][type='text']")
    # password_input_locator = (By.CSS_SELECTOR, "input.sbms-textbox[name='password'][type='password']")
    # enter_btn_locator = (By.CSS_SELECTOR, "button.sbms-button-ex")
    #
    # login_input_area = wait.until(EC.element_to_be_clickable(login_input_locator))
    # login_input_area.send_keys(LOGIN)
    #
    # password_input_area = wait.until(EC.element_to_be_clickable(password_input_locator))
    # password_input_area.send_keys(PASSWORD)
    #
    # enter_btn = wait.until(EC.element_to_be_clickable(enter_btn_locator))
    # enter_btn.click()
    #
    # log_step('Вход в SBMS выполнен')
    #
    # wait.until(
    #     lambda driver: "overflow: hidden; direction: ltr;" in driver.find_element(By.TAG_NAME,
    #                                                                               "body").get_attribute(
    #         "style"))
    #
    # cabs_locator = (By.XPATH, '//a[@class="menu__a-vertical" and text()="Витрины"]')
    # subs_cabinet_locator = (By.XPATH, '//a[@class="menu__a-vertical" and text()="Витрина абонента"]')
    #
    # cabs = wait.until(EC.element_to_be_clickable(cabs_locator))
    # cabs.click()
    #
    # subs_cabinet = wait.until(EC.element_to_be_clickable(subs_cabinet_locator))
    # subs_cabinet.click()
    #
    # wait.until(
    #     lambda driver: "overflow: hidden; direction: ltr;" in driver.find_element(By.TAG_NAME, "body").get_attribute(
    #         "style"))

    log_step('Витрина абонента открыта')

    input_tel_number_locator = (By.CLASS_NAME, "inp-text")
    input_tel_number = wait.until(EC.element_to_be_clickable(input_tel_number_locator))
    input_tel_number.send_keys("998501041717")

    log_step('MSISDN введен')

    search_elements_locator = (By.CSS_SELECTOR, 'ps-icon[icon="search-white"]')
    search_elements = wait.until(EC.element_to_be_clickable(search_elements_locator))
    search_elements.click()

    approve_num_btn_locator = (By.XPATH, "//span[@class='b-button__label' and text()='Да']/..")
    approve_num_btn = wait.until(EC.element_to_be_clickable(approve_num_btn_locator))
    approve_num_btn.click()

    log_step('Поиск выполнен и номер подтвержден')

    balance_of_subs_before = (By.CSS_SELECTOR, 'span#ps_customer_subscriber_summary_common_balance')
    wait.until(EC.element_to_be_clickable(balance_of_subs_before))
    balance_before_activating_rate_plan = driver.find_element(By.CSS_SELECTOR,
                                                              'span#ps_customer_subscriber_summary_common_balance')
    balance_before_activating_rate_plan_text = \
    balance_before_activating_rate_plan.text.replace('UZS', '').strip().replace(' ', '').split('.')[0]
    log_step(f'Баланс до подключения ТП: {int(balance_before_activating_rate_plan_text)}')

    clients_btn_locator = (By.XPATH, '//a[@class="menu__a-vertical" and text()="Клиенты"]')
    clients_btn = wait.until(EC.element_to_be_clickable(clients_btn_locator))
    clients_btn.click()

    abonents_btn_locator = (By.XPATH, '//a[@class="menu__a-vertical" and text()="Абоненты"]')
    abonents_btn = wait.until(EC.element_to_be_clickable(abonents_btn_locator))
    abonents_btn.click()

    services_cathegory_btn = (By.XPATH, '//a[@class="menu__a-vertical" and text()="Карточка абонента"]')
    services_cathegory_btn = wait.until(EC.element_to_be_clickable(services_cathegory_btn))
    services_cathegory_btn.click()
    time.sleep(5)

    log_step('Перешли в карточку абонента')

    driver.switch_to.frame('fr3')

    if driver.find_element(By.ID, 'RTPL').text == RATE_PLAN_NAME_2:
        try:
            change_rtpl_btn = wait.until(
                EC.element_to_be_clickable((By.ID, "BTN_CHANGE_RTPL"))
            )
            change_rtpl_btn.click()
            time.sleep(2)
        finally:
            driver.switch_to.default_content()

        input_field = wait.until(EC.visibility_of_element_located((By.XPATH,
                                                                   "//input[contains(@class, 'b-combobox__input') and @ng-readonly='states.readonly || states.disabled' and @ng-required='states.required']")))

        input_field.send_keys(RATE_PLAN_NAME_1)
        time.sleep(2)

        change_rtpl_checkbox = wait.until(
            EC.element_to_be_clickable((By.XPATH,
                                        "//div[@class='n-grid__text' and @data-row='$view.rows[0]' and @data-cell='$view.columns[0].getCell($view.rows[0])']/span[@class='n-check-checkbox']"))
        )
        change_rtpl_checkbox.click()
        log_step(f'Выбран {RATE_PLAN_NAME_1}')
        time.sleep(3)

        change_rtpl_btn_final_locator = (By.XPATH,
                                         "//ps-button[@ng-click='change()' and @class='b-button ps-component']/span[contains(@class, 'b-button__label') and text()='Сменить ТП']")

        change_rtpl_btn_final = wait.until(EC.element_to_be_clickable(change_rtpl_btn_final_locator))
        change_rtpl_btn_final.click()

        change_rtpl_yes_locator = (By.XPATH,
                                   '//ps-button[@ng-repeat="button in config.buttons track by $index" and @ng-click="psDialog.close(button.result)" and @class="b-button ps-component"]/span[contains(@class, "b-button__label") and text()="Да"]')
        change_rtpl_yes = wait.until(EC.element_to_be_clickable(change_rtpl_yes_locator))
        change_rtpl_yes.click()
        log_step(f'Смена ТП с {RATE_PLAN_NAME_2} на {RATE_PLAN_NAME_1} успешно выполнена!')
        result_message = f'Было {RATE_PLAN_NAME_2}, Стало: {RATE_PLAN_NAME_1}'
        time.sleep(2)
    else:
        try:
            change_rtpl_btn = wait.until(
                EC.element_to_be_clickable((By.ID, "BTN_CHANGE_RTPL"))
            )
            change_rtpl_btn.click()
            time.sleep(2)
        finally:
            driver.switch_to.default_content()

        input_field_locator = (By.XPATH,
                               "//input[contains(@class, 'b-combobox__input') and @ng-readonly='states.readonly || states.disabled' and @ng-required='states.required']")

        input_field = wait.until(EC.visibility_of_element_located(input_field_locator))

        input_field.send_keys(RATE_PLAN_NAME_2)
        time.sleep(2)

        change_rtpl_checkbox_locator = (By.XPATH,
                                        "//div[@class='n-grid__text' and @data-row='$view.rows[0]' and @data-cell='$view.columns[0].getCell($view.rows[0])']/span[@class='n-check-checkbox']")

        change_rtpl_checkbox = wait.until(
            EC.element_to_be_clickable(change_rtpl_checkbox_locator)
        )
        change_rtpl_checkbox.click()
        log_step(f'Выбран {RATE_PLAN_NAME_2}')
        time.sleep(3)

        change_rtpl_btn_final_locator = (By.XPATH,
                                         "//ps-button[@ng-click='change()' and @class='b-button ps-component']/span[contains(@class, 'b-button__label') and text()='Сменить ТП']")

        change_rtpl_btn_final = wait.until(EC.element_to_be_clickable(change_rtpl_btn_final_locator))
        change_rtpl_btn_final.click()

        change_rtpl_yes_locator = (By.XPATH,
                                   '//ps-button[@ng-repeat="button in config.buttons track by $index" and @ng-click="psDialog.close(button.result)" and @class="b-button ps-component"]/span[contains(@class, "b-button__label") and text()="Да"]')

        change_rtpl_yes = wait.until(EC.element_to_be_clickable(change_rtpl_yes_locator))
        change_rtpl_yes.click()
        log_step(f'Смена ТП с {RATE_PLAN_NAME_1} на {RATE_PLAN_NAME_2} успешно выполнена!')
        result_message = f'Было {RATE_PLAN_NAME_1}, Стало: {RATE_PLAN_NAME_2}'
        time.sleep(2)

    # Нажать обновить баланс 3 раза в течение 30 секунды
    refresh_btn = driver.find_element(By.XPATH, '//ps-button[@ng-click="updateBalances();"]')
    for _ in range(3):
        refresh_btn.click()
        time.sleep(10)

    # Получение баланса после смены
    balance_of_subs_after = (By.CSS_SELECTOR, 'span#ps_customer_subscriber_summary_common_balance')
    balance_after_activating_rate_plan = wait.until(EC.visibility_of_element_located(balance_of_subs_after))
    balance_after_activating_rate_plan_text = \
    balance_after_activating_rate_plan.text.replace('UZS', '').strip().replace(' ', '').split('.')[0]
    log_step(f'Баланс после подключения ТП: {int(balance_after_activating_rate_plan_text)}')

    # Записываем данные в Excel
    new_row = ['СМЕНА ТП',
               int(balance_before_activating_rate_plan_text),
               int(balance_after_activating_rate_plan_text),
               result_message]

    ws.append(new_row)
    wb.save(EXCEL_FILE)
