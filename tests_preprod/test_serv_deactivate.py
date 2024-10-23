from openpyxl import Workbook
import time
from datetime import datetime
from openpyxl.reader.excel import load_workbook

from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

from utils.login_x_password import LOGIN, PASSWORD
from config.config import PHONE_NUM

LOG_FILE = "../results/logs.txt"
EXCEL_FILE = "../results/SBMS_AUTOTEST_RESULTS.xlsx"

def log_step(message):
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_message = f"[{current_time}] {message}\n"
    print(log_message.strip())
    with open(LOG_FILE, "a") as file:
        file.write(log_message)

def test_serv_deactivate(driver):
    global wb, ws

    # Попробуйте открыть существующий файл или создать новый
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active

    driver.get("https://sbms.ucell/ps/sbms/shell.html")
    wait = WebDriverWait(driver, 120)

    log_step(' ======== Отключение Услуги ========')

    # # Проверка кнопки
    # detail_button_check = wait.until(
    #     EC.visibility_of_element_located((By.ID, "details-button"))
    # )
    # detail_button_check.click()
    #
    # go_to_link = wait.until(
    #     EC.visibility_of_element_located((By.ID, "proceed-link"))
    # )
    # go_to_link.click()
    # login_text_locator = (By.CSS_SELECTOR, '.login-caption > span')
    # login_locator_check = wait.until(EC.visibility_of_element_located(login_text_locator))
    # assert login_locator_check.is_displayed(), "Не найден локатор входа"
    #
    # login_input_locator = (By.CSS_SELECTOR, "input.sbms-textbox[name='user'][type='text']")
    # password_input_locator = (By.CSS_SELECTOR, "input.sbms-textbox[name='password'][type='password']")
    # enter_btn_locator = (By.CSS_SELECTOR, "button.sbms-button-ex")
    #
    # # Ожидаем, пока поле ввода логина будет доступно для ввода
    # login_input_area = wait.until(EC.element_to_be_clickable(login_input_locator))
    # login_input_area.send_keys(LOGIN)
    #
    # # Ожидаем, пока поле ввода пароля будет доступно для ввода
    # password_input_area = wait.until(EC.element_to_be_clickable(password_input_locator))
    # password_input_area.send_keys(PASSWORD)
    #
    # # Ожидаем, пока кнопка "Войти" станет кликабельной
    # enter_btn = wait.until(EC.element_to_be_clickable(enter_btn_locator))
    # enter_btn.click()

    log_step('Вход в SBMS выполнен')

    # Ждать пока страничка не загрузиться полностью
    wait.until(
        lambda driver: "overflow: hidden; direction: ltr;" in driver.find_element(By.TAG_NAME,
                                                                                  "body").get_attribute(
            "style"))

    # # Вход в витрину
    #
    # Локаторы для элементов меню
    cabs_locator = (By.XPATH, '//a[@class="menu__a-vertical" and text()="Витрины"]')
    subs_cabinet_locator = (By.XPATH, '//a[@class="menu__a-vertical" and text()="Витрина абонента"]')

    # Ожидание кликабельности элемента "Витрины" и клик по нему
    cabs = wait.until(EC.element_to_be_clickable(cabs_locator))
    cabs.click()

    # Ожидание кликабельности элемента "Витрина абонента" и клик по нему
    subs_cabinet = wait.until(EC.element_to_be_clickable(subs_cabinet_locator))
    subs_cabinet.click()

    wait.until(
        lambda driver: "overflow: hidden; direction: ltr;" in driver.find_element(By.TAG_NAME, "body").get_attribute(
            "style"))

    log_step('Витрина абонента открыта')

    # Ввод MSISDN
    input_tel_number_locator = (By.CLASS_NAME, "inp-text")
    input_tel_number = wait.until(EC.element_to_be_clickable(input_tel_number_locator))
    input_tel_number.send_keys(PHONE_NUM)
    # time.sleep(5)

    log_step('MSISDN введен')

    # Клик по кнопке ПОИСК
    search_elements_locator = (By.CSS_SELECTOR, 'ps-icon[icon="search-white"]')
    search_elements = wait.until(EC.element_to_be_clickable(search_elements_locator))
    search_elements.click()
    # time.sleep(5)

    approve_num_btn_locator = (By.XPATH, "//span[@class='b-button__label' and text()='Да']/..")
    approve_num_btn = wait.until(EC.element_to_be_clickable(approve_num_btn_locator))
    approve_num_btn.click()
    # time.sleep(5)

    log_step('Поиск выполнен и номер подтвержден')

    balance_of_subs_before = (By.CSS_SELECTOR, 'span#ps_customer_subscriber_summary_common_balance')
    wait.until(EC.element_to_be_clickable(balance_of_subs_before))
    balance_before_deactivating_service = driver.find_element(By.CSS_SELECTOR,
                                                              'span#ps_customer_subscriber_summary_common_balance')
    balance_before_deactivating_service_text = \
        balance_before_deactivating_service.text.replace('UZS', '').strip().replace(' ', '').split('.')[0]
    log_step(f'Баланс до подключения ТП: {int(balance_before_deactivating_service_text)}')

    # ============================= ТЕСТИРОВАНИЕ УСЛУГИ НАЧИНАЕТСЯ ЗДЕСЬ =============================

    clients_btn_locator = (By.XPATH, '//a[@class="menu__a-vertical" and text()="Клиенты"]')
    clients_btn = wait.until(EC.element_to_be_clickable(clients_btn_locator))
    clients_btn.click()

    abonents_btn_locator = (By.XPATH, '//a[@class="menu__a-vertical" and text()="Абоненты"]')
    abonents_btn = wait.until(EC.element_to_be_clickable(abonents_btn_locator))
    abonents_btn.click()

    services_cathegory_btn_locator = (By.XPATH, '//a[@class="menu__a-vertical" and text()="Услуги"]')
    services_cathegory_btn = wait.until(EC.element_to_be_clickable(services_cathegory_btn_locator))
    services_cathegory_btn.click()

    log_step('Перешли в раздел Услуги')

    """ 
    ============================================= РАЗДЕЛ УСЛУГИ ==================================================
    """

    packs_btn_locator = (By.XPATH, '//a[@class="n-tab__title" and text()="Пакеты"]')
    packs_btn = wait.until(EC.element_to_be_clickable(packs_btn_locator))
    packs_btn.click()

    """ ============================================= HERE GOES DEACTIVATING THE SERVICE ============================================= """

    serv_search_input_clear_locator = (By.XPATH,
                                                  "//input[@ng-model='grdPacks.filter.name' and contains(@class, 'inp-text')]")
    serv_search_input_clear = wait.until(EC.element_to_be_clickable(serv_search_input_clear_locator))
    serv_search_input_clear.send_keys(Keys.COMMAND + "A") # Для Mac Keys.COMMAND
    serv_search_input_clear.send_keys(Keys.DELETE)
    log_step('Очищен поле поиска Услуги')

    serv_search_input_locator = (By.XPATH,
                                            "//input[@ng-model='grdPacks.filter.name' and contains(@class, 'inp-text')]")
    serv_search_input = wait.until(EC.element_to_be_clickable(serv_search_input_locator))
    serv_search_input.send_keys('Traffic')
    log_step('Поиск услуги Traffic*')

    # Locate the element
    serv_state_click_locator = (By.XPATH, '//div[@class="n-grid-title-in" and @style="z-index: 8;"]')
    serv_state_click = wait.until(EC.element_to_be_clickable(serv_state_click_locator))

    # Perform a double-click
    actions = ActionChains(driver)
    actions.double_click(serv_state_click).perform()

    time.sleep(2)
    serv_search_input.click()

    delete_serv_btn_locator = (By.CSS_SELECTOR,
                                          'ps-button[title="Отключить пакет"] > ps-icon[icon="row-delete"]')
    delete_serv_btn = wait.until(EC.element_to_be_clickable(delete_serv_btn_locator))
    delete_serv_btn.click()
    time.sleep(2)

    serv_comment_input_locator = (By.XPATH, '//input[@ng-model="deactivateServices.deleteComment"]')
    serv_comment_input = wait.until(EC.element_to_be_clickable(serv_comment_input_locator))
    serv_comment_input.send_keys('TEST')
    time.sleep(2)

    serv_deactivate_btn_locator = (By.XPATH, '//span[@class="b-button__label" and text()="Отключить"]')
    serv_deactivate_btn = wait.until(EC.element_to_be_clickable(serv_deactivate_btn_locator))
    serv_deactivate_btn.click()
    time.sleep(2)
    log_step('Услуга отключена!')

    """ ============================================= Обновим баланс ============================================= """

    # Нажать обновить баланс 3 раза в течение 30 секунды
    refresh_btn_locator = (By.XPATH, '//ps-button[@ng-click="updateBalances();"]')
    refresh_btn = wait.until(EC.element_to_be_clickable(refresh_btn_locator))
    for _ in range(3):
        refresh_btn.click()
        time.sleep(15)
    log_step('Кнопка обновить был нажат 3 раза в течение 30 секунды')

    balance_of_subs_after = (By.CSS_SELECTOR, 'span#ps_customer_subscriber_summary_common_balance')
    wait.until(EC.element_to_be_clickable(balance_of_subs_after))
    balance_after_deactivating_service = driver.find_element(By.CSS_SELECTOR,
                                                             'span#ps_customer_subscriber_summary_common_balance')
    balance_after_deactivating_service_text = \
        balance_after_deactivating_service.text.replace('UZS', '').strip().replace(' ', '').split('.')[0]
    log_step(f'Баланс после подключения ТП: {int(balance_after_deactivating_service_text)}')

    time.sleep(2)
    log_step('Успешно отключен пакет Traffic+!')

    # Записываем данные в Excel
    new_row = ['Отключение услуги',
               int(balance_before_deactivating_service_text),
               int(balance_after_deactivating_service_text),
               'Отключена услуга Traffic+',
               datetime.now().strftime("%Y-%m-%d %H:%M:%S")
               ]

    ws.append(new_row)
    wb.save(EXCEL_FILE)