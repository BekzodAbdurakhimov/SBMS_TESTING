from openpyxl import Workbook
import time
from datetime import datetime
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from trio import current_time

from config import LOGIN, PASSWORD, RATE_PLAN_NAME_1, RATE_PLAN_NAME_2
import pytest

LOG_FILE = "logs.txt"

def log_step(message):
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_message = f"[{current_time}] {message}\n"
    print(log_message.strip())
    with open(LOG_FILE, "a") as file:
        file.write(log_message)

def test_rate_plan(driver):
    global wb, ws
    wb = Workbook()
    ws = wb.active

    driver.get("https://sbms.ucell/ps/sbms/shell.html")
    wait = WebDriverWait(driver, 120)

    log_step(' ======== Смена ТП ========')

    # Проверка кнопки
    detail_button_check = wait.until(
        EC.visibility_of_element_located((By.ID, "details-button"))
    )
    detail_button_check.click()

    go_to_link = wait.until(
        EC.visibility_of_element_located((By.ID, "proceed-link"))
    )
    go_to_link.click()
    login_text_locator = (By.CSS_SELECTOR, '.login-caption > span')
    login_locator_check = wait.until(EC.visibility_of_element_located(login_text_locator))
    assert login_locator_check.is_displayed(), "Не найден локатор входа"

    # time.sleep(5)

    login_input_locator = (By.CSS_SELECTOR, "input.sbms-textbox[name='user'][type='text']")
    password_input_locator = (By.CSS_SELECTOR, "input.sbms-textbox[name='password'][type='password']")
    enter_btn_locator = (By.CSS_SELECTOR, "button.sbms-button-ex")

    # Ожидаем, пока поле ввода логина будет доступно для ввода
    login_input_area = wait.until(EC.element_to_be_clickable(login_input_locator))
    login_input_area.send_keys(LOGIN)

    # Ожидаем, пока поле ввода пароля будет доступно для ввода
    password_input_area = wait.until(EC.element_to_be_clickable(password_input_locator))
    password_input_area.send_keys(PASSWORD)

    # Ожидаем, пока кнопка "Войти" станет кликабельной
    enter_btn = wait.until(EC.element_to_be_clickable(enter_btn_locator))
    enter_btn.click()

    log_step('Вход в SBMS выполнен')

    # Ждать пока страничка не загрузиться полностью
    wait.until(
        lambda driver: "overflow: hidden; direction: ltr;" in driver.find_element(By.TAG_NAME,
                                                                                  "body").get_attribute(
            "style"))


    # Вход в витрину

    # Локаторы для элементов меню
    cabs_locator = (By.XPATH, '//a[@class="menu__a-vertical" and text()="Витрины"]')
    subs_cabinet_locator = (By.XPATH, '//a[@class="menu__a-vertical" and text()="Витрина абонента"]')

    # Ожидание кликабельности элемента "Витрины" и клик по нему
    cabs = wait.until(EC.element_to_be_clickable(cabs_locator))
    cabs.click()

    # Ожидание кликабельности элемента "Витрина абонента" и клик по нему
    subs_cabinet = wait.until(EC.element_to_be_clickable(subs_cabinet_locator))
    subs_cabinet.click()

    wait.until(
        lambda driver: "overflow: hidden; direction: ltr;" in driver.find_element(By.TAG_NAME, "body").get_attribute(
            "style"))

    log_step('Витрина абонента открыта')

    # Ввод MSISDN
    input_tel_number_locator = (By.CLASS_NAME, "inp-text")
    input_tel_number = wait.until(EC.element_to_be_clickable(input_tel_number_locator))
    input_tel_number.send_keys("998501041717")
    # time.sleep(5)

    log_step('MSISDN введен')

    # Клик по кнопке ПОИСК
    search_elements_locator = (By.CSS_SELECTOR, 'ps-icon[icon="search-white"]')
    search_elements = wait.until(EC.element_to_be_clickable(search_elements_locator))
    search_elements.click()
    # time.sleep(5)

    approve_num_btn_locator = (By.XPATH, "//span[@class='b-button__label' and text()='Да']/..")
    approve_num_btn = wait.until(EC.element_to_be_clickable(approve_num_btn_locator))
    approve_num_btn.click()
    # time.sleep(5)

    log_step('Поиск выполнен и номер подтвержден')

    balance_of_subs_before = (By.CSS_SELECTOR, 'span#ps_customer_subscriber_summary_common_balance')
    wait.until(EC.element_to_be_clickable(balance_of_subs_before))
    balance_before_activating_rate_plan = driver.find_element(By.CSS_SELECTOR, 'span#ps_customer_subscriber_summary_common_balance')
    balance_before_activating_rate_plan_text = balance_before_activating_rate_plan.text.replace('UZS', '').strip().replace(' ', '').split('.')[0]
    log_step(f'Баланс до подключения ТП: {int(balance_before_activating_rate_plan_text)}')
    time.sleep(2)
# ============================= ТЕСТИРОВАНИЕ ТП НАЧИНАЕТСЯ ЗДЕСЬ =============================

# ============================= ТЕСТИРОВАНИЕ ТП НАЧИНАЕТСЯ ЗДЕСЬ =============================

    clients_btn_locator = (By.XPATH, '//a[@class="menu__a-vertical" and text()="Клиенты"]')
    clients_btn = wait.until(EC.element_to_be_clickable(clients_btn_locator))
    clients_btn.click()
    # time.sleep(2)

    abonents_btn_locator = (By.XPATH, '//a[@class="menu__a-vertical" and text()="Абоненты"]')
    abonents_btn = wait.until(EC.element_to_be_clickable(abonents_btn_locator))
    abonents_btn.click()
    # time.sleep(2)

    services_cathegory_btn = (By.XPATH, '//a[@class="menu__a-vertical" and text()="Карточка абонента"]')
    services_cathegory_btn = wait.until(EC.element_to_be_clickable(services_cathegory_btn))
    services_cathegory_btn.click()
    time.sleep(5)

    log_step('Перешли в карточку абонента')

    # Переключение на iframe ТП
    driver.switch_to.frame('fr3')

    if driver.find_element(By.ID, 'RTPL').text == RATE_PLAN_NAME_2:
        try:
            change_rtpl_btn = wait.until(
                EC.element_to_be_clickable((By.ID, "BTN_CHANGE_RTPL"))
            )
            change_rtpl_btn.click()
            time.sleep(2)
        finally:
            driver.switch_to.default_content()

        input_field = wait.until(EC.visibility_of_element_located((By.XPATH,
                                                                   "//input[contains(@class, 'b-combobox__input') and @ng-readonly='states.readonly || states.disabled' and @ng-required='states.required']")))
        input_field.send_keys(RATE_PLAN_NAME_1)
        time.sleep(2)

        change_rtpl_checkbox = wait.until(
            EC.element_to_be_clickable((By.XPATH,
                                        "//div[@class='n-grid__text' and @data-row='$view.rows[0]' and @data-cell='$view.columns[0].getCell($view.rows[0])']/span[@class='n-check-checkbox']"))
        )
        change_rtpl_checkbox.click()
        log_step(f'Выбран {RATE_PLAN_NAME_1}')
        time.sleep(3)

        change_rtpl_btn_final_locator = (By.XPATH,
                                                    "//ps-button[@ng-click='change()' and @class='b-button ps-component']/span[contains(@class, 'b-button__label') and text()='Сменить ТП']")
        change_rtpl_btn_final = wait.until(EC.element_to_be_clickable(change_rtpl_btn_final_locator))
        change_rtpl_btn_final.click()
        # time.sleep(2)

        change_rtpl_yes_locator = (By.XPATH,
                                              '//ps-button[@ng-repeat="button in config.buttons track by $index" and @ng-click="psDialog.close(button.result)" and @class="b-button ps-component"]/span[contains(@class, "b-button__label") and text()="Да"]')
        change_rtpl_yes = wait.until(EC.element_to_be_clickable(change_rtpl_yes_locator))
        change_rtpl_yes.click()
        log_step(f'Смена ТП с {RATE_PLAN_NAME_2} на {RATE_PLAN_NAME_1} успешно выполнена!')
        ws['D1'] = f'Было {RATE_PLAN_NAME_2}, Стало: {RATE_PLAN_NAME_1}'
        time.sleep(2)
    else:
        try:
            change_rtpl_btn = wait.until(
                EC.element_to_be_clickable((By.ID, "BTN_CHANGE_RTPL"))
            )
            change_rtpl_btn.click()
            time.sleep(2)
        finally:
            driver.switch_to.default_content()

        input_field_locator = (By.XPATH,
                                          "//input[contains(@class, 'b-combobox__input') and @ng-readonly='states.readonly || states.disabled' and @ng-required='states.required']")
        input_field = wait.until(EC.element_to_be_clickable(input_field_locator))
        input_field.send_keys(RATE_PLAN_NAME_2)
        # time.sleep(2)

        change_rtpl_checkbox = wait.until(
            EC.element_to_be_clickable((By.XPATH,
                                        "//div[@class='n-grid__text' and @data-row='$view.rows[0]' and @data-cell='$view.columns[0].getCell($view.rows[0])']/span[@class='n-check-checkbox']"))
        )
        change_rtpl_checkbox.click()
        log_step(f'Выбран {RATE_PLAN_NAME_2}')
        # time.sleep(2)

        change_rtpl_btn_final_locator = (By.XPATH,
                                                    "//ps-button[@ng-click='change()' and @class='b-button ps-component']/span[contains(@class, 'b-button__label') and text()='Сменить ТП']")
        change_rtpl_btn_final = wait.until(EC.element_to_be_clickable(change_rtpl_btn_final_locator))
        change_rtpl_btn_final.click()
        # time.sleep(5)

        # change_rtpl_yes = driver.find_element(By.XPATH,
        #                                       '//ps-button[@ng-repeat="button in config.buttons track by $index" and @ng-click="psDialog.close(button.result)" and @class="b-button ps-component"]/span[contains(@class, "b-button__label") and text()="Да"]')
        # change_rtpl_yes.click()

        change_rtpl_yes = wait.until(
            EC.element_to_be_clickable((By.XPATH,
                                        '//ps-button[@ng-repeat="button in config.buttons track by $index" and @ng-click="psDialog.close(button.result)" and @class="b-button ps-component"]/span[contains(@class, "b-button__label") and text()="Да"]'))
        )
        change_rtpl_yes.click()

        log_step(f'Смена ТП с {RATE_PLAN_NAME_1} на {RATE_PLAN_NAME_2} успешно выполнена!')
        ws['D1'] = f'Было {RATE_PLAN_NAME_1}, Стало: {RATE_PLAN_NAME_2}'
        time.sleep(2)

    # Нажать обновить баланс 3 раза в течение 30 секунды
    refresh_btn = driver.find_element(By.XPATH, '//ps-button[@ng-click="updateBalances();"]')
    for _ in range(3):
        refresh_btn.click()
        time.sleep(10)

    balance_of_subs_after = (By.CSS_SELECTOR, 'span#ps_customer_subscriber_summary_common_balance')
    wait.until(EC.element_to_be_clickable(balance_of_subs_after))
    balance_after_activating_rate_plan = driver.find_element(By.CSS_SELECTOR,
                                                                 'span#ps_customer_subscriber_summary_common_balance')
    balance_after_activating_rate_plan_text = \
    balance_after_activating_rate_plan.text.replace('UZS', '').strip().replace(' ', '').split('.')[0]
    log_step(f'Баланс после подключения ТП: {int(balance_after_activating_rate_plan_text)}')

    time.sleep(2)

    ws['A1'] = "СМЕНА ТП"
    ws['B1'] = int(balance_before_activating_rate_plan_text)
    ws['C1'] = int(balance_after_activating_rate_plan_text)
    ws['E1'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    wb.save("SBMS_AUTOTEST_RESULTS.xlsx")
